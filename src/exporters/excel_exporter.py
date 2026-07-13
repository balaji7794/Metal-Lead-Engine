from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelExporter:

    def export(self, leads, filename="output/leads.xlsx"):

        Path("output").mkdir(exist_ok=True)

        wb = Workbook()

        ws = wb.active

        ws.title = "Google Maps Leads"

        headers = [
            "Company Name",
            "Category",
            "Address",
            "City",
            "State",
            "Phone",
            "Email",
            "Website",
            "Rating",
            "Reviews",
            "Google Maps URL",
            "Source",
            "Notes"
        ]

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for lead in leads:

            ws.cell(row=row, column=1).value = lead.name
            ws.cell(row=row, column=2).value = lead.category
            ws.cell(row=row, column=3).value = lead.address
            ws.cell(row=row, column=4).value = lead.city
            ws.cell(row=row, column=5).value = lead.state
            ws.cell(row=row, column=6).value = lead.phone
            ws.cell(row=row, column=7).value = ", ".join(lead.emails)
            ws.cell(row=row, column=8).value = lead.website
            ws.cell(row=row, column=9).value = lead.rating
            ws.cell(row=row, column=10).value = lead.review_count
            ws.cell(row=row, column=11).value = lead.google_maps_url
            ws.cell(row=row, column=12).value = lead.source
            ws.cell(row=row, column=13).value = lead.notes

            row += 1

        wb.save(filename)

        print(f"\n✅ Excel exported : {filename}")