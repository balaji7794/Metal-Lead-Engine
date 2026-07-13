from pathlib import Path

from openpyxl import load_workbook


FILE = Path(
    "knowledge/master_data/aluminium_raw_materials.xlsx"
)


RAW_MATERIALS = [

# ==========================================================
# PRIMARY ALUMINIUM
# ==========================================================

{"id":"RM0001","name":"Primary Aluminium Ingot","family":"Primary Aluminium","category":"Primary","form":"Ingot"},
{"id":"RM0002","name":"Primary Aluminium Billet","family":"Primary Aluminium","category":"Primary","form":"Billet"},
{"id":"RM0003","name":"Primary Aluminium Slab","family":"Primary Aluminium","category":"Primary","form":"Slab"},
{"id":"RM0004","name":"Primary Aluminium Sow","family":"Primary Aluminium","category":"Primary","form":"Sow"},
{"id":"RM0005","name":"Primary Aluminium Wire Rod","family":"Primary Aluminium","category":"Primary","form":"Wire Rod"},
{"id":"RM0006","name":"Primary Aluminium Liquid Metal","family":"Primary Aluminium","category":"Primary","form":"Liquid"},

# ==========================================================
# SECONDARY ALUMINIUM
# ==========================================================

{"id":"RM0101","name":"Secondary Aluminium Remelt Ingot","family":"Secondary","category":"Secondary","form":"Ingot"},
{"id":"RM0102","name":"Secondary Aluminium Alloy Ingot","family":"Secondary","category":"Secondary","form":"Ingot"},
{"id":"RM0103","name":"Secondary Aluminium Billet","family":"Secondary","category":"Secondary","form":"Billet"},
{"id":"RM0104","name":"Secondary Aluminium Sow","family":"Secondary","category":"Secondary","form":"Sow"},
{"id":"RM0105","name":"Secondary Aluminium Liquid Metal","family":"Secondary","category":"Secondary","form":"Liquid"},

# ==========================================================
# BILLETS
# ==========================================================

{"id":"RM0201","name":"1050 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0202","name":"1060 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0203","name":"1070 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0204","name":"1100 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0205","name":"2014 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0206","name":"2024 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0207","name":"3003 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0208","name":"3105 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0209","name":"5005 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0210","name":"5052 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0211","name":"5083 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0212","name":"5086 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0213","name":"5154 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0214","name":"5251 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0215","name":"5454 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0216","name":"5754 Billet","family":"Billet","category":"Primary","form":"Billet"},
{"id":"RM0217","name":"6060 Billet","family":"Billet","category":"Extrusion","form":"Billet"},
{"id":"RM0218","name":"6061 Billet","family":"Billet","category":"Extrusion","form":"Billet"},
{"id":"RM0219","name":"6063 Billet","family":"Billet","category":"Extrusion","form":"Billet"},
{"id":"RM0220","name":"6082 Billet","family":"Billet","category":"Extrusion","form":"Billet"},
{"id":"RM0221","name":"6101 Billet","family":"Billet","category":"Extrusion","form":"Billet"},
{"id":"RM0222","name":"6262 Billet","family":"Billet","category":"Extrusion","form":"Billet"},
{"id":"RM0223","name":"6351 Billet","family":"Billet","category":"Extrusion","form":"Billet"},
{"id":"RM0224","name":"6463 Billet","family":"Billet","category":"Extrusion","form":"Billet"},
{"id":"RM0225","name":"7005 Billet","family":"Billet","category":"High Strength","form":"Billet"},
{"id":"RM0226","name":"7020 Billet","family":"Billet","category":"High Strength","form":"Billet"},
{"id":"RM0227","name":"7075 Billet","family":"Billet","category":"High Strength","form":"Billet"},

# ==========================================================
# ALLOY INGOTS
# ==========================================================

{"id":"RM0301","name":"ADC1 Alloy Ingot","family":"ADC","category":"Die Casting","form":"Ingot"},
{"id":"RM0302","name":"ADC3 Alloy Ingot","family":"ADC","category":"Die Casting","form":"Ingot"},
{"id":"RM0303","name":"ADC5 Alloy Ingot","family":"ADC","category":"Die Casting","form":"Ingot"},
{"id":"RM0304","name":"ADC6 Alloy Ingot","family":"ADC","category":"Die Casting","form":"Ingot"},
{"id":"RM0305","name":"ADC10 Alloy Ingot","family":"ADC","category":"Die Casting","form":"Ingot"},
{"id":"RM0306","name":"ADC12 Alloy Ingot","family":"ADC","category":"Die Casting","form":"Ingot"},
{"id":"RM0307","name":"ADC14 Alloy Ingot","family":"ADC","category":"Die Casting","form":"Ingot"},

{"id":"RM0310","name":"LM2 Alloy Ingot","family":"LM","category":"Foundry","form":"Ingot"},
{"id":"RM0311","name":"LM4 Alloy Ingot","family":"LM","category":"Foundry","form":"Ingot"},
{"id":"RM0312","name":"LM6 Alloy Ingot","family":"LM","category":"Foundry","form":"Ingot"},
{"id":"RM0313","name":"LM9 Alloy Ingot","family":"LM","category":"Foundry","form":"Ingot"},
{"id":"RM0314","name":"LM13 Alloy Ingot","family":"LM","category":"Foundry","form":"Ingot"},
{"id":"RM0315","name":"LM24 Alloy Ingot","family":"LM","category":"Foundry","form":"Ingot"},
{"id":"RM0316","name":"LM25 Alloy Ingot","family":"LM","category":"Foundry","form":"Ingot"},

# ==========================================================
# MASTER ALLOYS
# ==========================================================

{"id":"RM0401","name":"Aluminium Silicon Master Alloy","family":"Master Alloy","category":"Master Alloy","form":"Ingot"},
{"id":"RM0402","name":"Aluminium Titanium Master Alloy","family":"Master Alloy","category":"Master Alloy","form":"Ingot"},
{"id":"RM0403","name":"Aluminium Boron Master Alloy","family":"Master Alloy","category":"Master Alloy","form":"Ingot"},
{"id":"RM0404","name":"Aluminium Strontium Master Alloy","family":"Master Alloy","category":"Master Alloy","form":"Ingot"},
{"id":"RM0405","name":"Aluminium Manganese Master Alloy","family":"Master Alloy","category":"Master Alloy","form":"Ingot"},
{"id":"RM0406","name":"Aluminium Chromium Master Alloy","family":"Master Alloy","category":"Master Alloy","form":"Ingot"},
{"id":"RM0407","name":"Aluminium Copper Master Alloy","family":"Master Alloy","category":"Master Alloy","form":"Ingot"},
{"id":"RM0408","name":"Aluminium Magnesium Master Alloy","family":"Master Alloy","category":"Master Alloy","form":"Ingot"},
{"id":"RM0409","name":"Aluminium Zinc Master Alloy","family":"Master Alloy","category":"Master Alloy","form":"Ingot"},
{"id":"RM0410","name":"Aluminium Nickel Master Alloy","family":"Master Alloy","category":"Master Alloy","form":"Ingot"},

# ==========================================================
# GRAIN REFINERS
# ==========================================================

{"id":"RM0501","name":"AlTi5B1","family":"Grain Refiner","category":"Refining","form":"Rod"},
{"id":"RM0502","name":"AlTi3B1","family":"Grain Refiner","category":"Refining","form":"Rod"},
{"id":"RM0503","name":"AlB3","family":"Grain Refiner","category":"Refining","form":"Rod"},
{"id":"RM0504","name":"AlB5","family":"Grain Refiner","category":"Refining","form":"Rod"},

# ==========================================================
# FLUXES
# ==========================================================

{"id":"RM0601","name":"Cover Flux","family":"Flux","category":"Melting","form":"Powder"},
{"id":"RM0602","name":"Cleaning Flux","family":"Flux","category":"Melting","form":"Powder"},
{"id":"RM0603","name":"Drossing Flux","family":"Flux","category":"Melting","form":"Powder"},
{"id":"RM0604","name":"Degassing Flux","family":"Flux","category":"Melting","form":"Powder"},
{"id":"RM0605","name":"Refining Flux","family":"Flux","category":"Melting","form":"Powder"},
{"id":"RM0606","name":"Slag Conditioner","family":"Flux","category":"Melting","form":"Powder"},

# ==========================================================
# DEGASSING
# ==========================================================

{"id":"RM0701","name":"Nitrogen Gas","family":"Degassing","category":"Gas","form":"Gas"},
{"id":"RM0702","name":"Argon Gas","family":"Degassing","category":"Gas","form":"Gas"},
{"id":"RM0703","name":"Chlorine Gas","family":"Degassing","category":"Gas","form":"Gas"},
{"id":"RM0704","name":"Nitrogen-Argon Mixture","family":"Degassing","category":"Gas","form":"Gas"}

]


def main():

    wb = load_workbook(FILE)

    ws = wb["Master Data"]

    # --------------------------------
    # Clear existing data (keep header)
    # --------------------------------

    if ws.max_row > 1:

        ws.delete_rows(2, ws.max_row - 1)

    for material in RAW_MATERIALS:

        ws.append([

            material["id"],
            material["name"],
            material["family"],
            material["category"],
            material["form"],

            "",      # Grade
            "",      # Produced By
            "",      # Used By
            "",      # Manufacturing Process
            "",      # Possible Finished Products
            "",      # Possible Scrap Generated
            "",      # Estimated Monthly Consumption
            100,     # Confidence
            "MetalEngine Knowledge Base",
            ""

        ])

    wb.save(FILE)

    print()
    print("=" * 60)
    print(f"Inserted {len(RAW_MATERIALS)} Raw Materials")
    print("=" * 60)


if __name__ == "__main__":

    main()